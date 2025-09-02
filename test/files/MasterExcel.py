import openpyxl

file_excel = openpyxl.load_workbook("C:\\Users\\lilia\\Desktop\\python-robot\\Test_Case\\files\\excel_python.xlsx")


def number_rows(hoja):
    ac =   file_excel[hoja]
    return ac.max_row

def data_columns(hoja, fila, columna):
    ac =   file_excel[hoja]
    columna= ac.cell(int(fila), int(columna))
    return columna.value


