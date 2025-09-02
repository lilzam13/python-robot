import openpyxl

base = openpyxl.load_workbook("C:\\Users\\lilia\\Desktop\\python-robot\\Test_Case\\files\\excel_python.xlsx")
print(base.sheetnames)
print("Pagina activa", base.active.title)

ac = base["Hoja1"]
filas = ac.max_row
columnas =  ac.max_column
esc =  ac["D2"].value="Pass"
esc =  ac["D3"].value="Pass"
esc =  ac["D4"].value="Error"

base.save("C:\\Users\\lilia\\Desktop\\python-robot\\Test_Case\\files\\excel_python.xlsx")

