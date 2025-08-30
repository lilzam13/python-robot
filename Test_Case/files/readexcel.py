import openpyxl

base = openpyxl.load_workbook("C:\\Users\\lilia\\Desktop\\python-robot\\Test_Case\\files\\excel_python.xlsx")

print(base.sheetnames)
print("Pagina activa", base.active.title)

ac = base["Hoja1"]
print(ac['A2'].value)
print(ac['B2'].value)
print(ac['C2'].value)


cel = ac.cell(1,2)
print(cel.value)