import openpyxl

base = openpyxl.load_workbook("C:\\Users\\lilia\\Desktop\\python-robot\\Test_Case\\files\\excel_python.xlsx")
print(base.sheetnames)
print("Pagina activa", base.active.title)

ac = base["Hoja1"]
filas = ac.max_row
columnas =  ac.max_column

print("Maximo de filas", filas)
print("Maximo de columnas", columnas)

"""for r  in ac['A2':'C4']:
    for c in r:
        print(c.value)"""

for r  in range(1, filas+1):
    for c in range(1, columnas+1):
        v = ac.cell(r,c)
        print(v.value)

